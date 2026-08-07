"""Builds the dynamic Excel workbook from a plugin's template_spec() and current config."""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL = PatternFill("solid", fgColor="1F3B57")
REFERENCE_FILL = PatternFill("solid", fgColor="6B7A8C")
REQUIRED_FILL = PatternFill("solid", fgColor="9C2B2B")
CONDITIONAL_FILL = PatternFill("solid", fgColor="8A5B00")
IGNORED_FILL = PatternFill("solid", fgColor="9AA5B1")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MAX_ROWS_PER_SHEET = 10_000

REQUIREMENT_FILLS = {
    "required": REQUIRED_FILL,
    "conditional": CONDITIONAL_FILL,
    "not supported": IGNORED_FILL,
}

# Rows beyond the exported data that still offer the dropdowns, so new entries are guided.
SPARE_ROWS = 500

# Cells the upload path reads back to tell which export a workbook came from.
PROVENANCE_LABELS = {"snapshot": "Snapshot ID", "exported": "Exported at"}

DROPDOWNS = {
    "action": '"create,update,delete"',
    "protocol": '"TCP,UDP"',
    "enabled": '"true,false"',
    "log_begin": '"true,false"',
    "log_end": '"true,false"',
    "send_events_to_fmc": '"true,false"',
    "rule_action": '"ALLOW,TRUST,BLOCK,MONITOR,BLOCK_RESET,BLOCK_INTERACTIVE,BLOCK_RESET_INTERACTIVE"',
}


def build_workbook(
    spec: dict[str, list[str]],
    product: str,
    product_version: str | None,
    existing: dict[str, list[dict[str, Any]]] | None = None,
    reference_sheets: set[str] | None = None,
    snapshot_id: str | None = None,
    snapshot_label: str | None = None,
    guide: dict[str, dict[str, tuple[str, str, str]]] | None = None,
) -> bytes:
    existing = existing or {}
    reference_sheets = reference_sheets or set()
    guide = guide or {}

    wb = Workbook()
    wb.remove(wb.active)
    _write_readme(
        wb, spec, existing, product, product_version, reference_sheets, snapshot_id, snapshot_label
    )
    if guide:
        _write_field_guide(wb, spec, guide, reference_sheets)

    for sheet_name, headers in spec.items():
        ws = wb.create_sheet(sheet_name[:31])
        ws.append(headers)

        is_reference = sheet_name in reference_sheets
        sheet_guide = guide.get(sheet_name, {})
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            requirement, description, example = sheet_guide.get(header, ("", "", ""))

            if is_reference:
                cell.fill = REFERENCE_FILL
            else:
                cell.fill = REQUIREMENT_FILLS.get(requirement, HEADER_FILL)
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = max(14, len(header) + 4)

            if description:
                note = f"{requirement.upper()}\n\n{description}"
                if example:
                    note += f"\n\nExample: {example}"
                cell.comment = Comment(note, "CSAP", height=170, width=320)

        for row in (existing.get(sheet_name) or [])[:MAX_ROWS_PER_SHEET]:
            ws.append([_cell(row.get(header)) for header in headers])

        _add_dropdowns(ws, headers)
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_field_guide(
    wb: Workbook,
    spec: dict[str, list[str]],
    guide: dict[str, dict[str, tuple[str, str, str]]],
    reference_sheets: set[str],
) -> None:
    """Every column on every sheet, so nobody has to guess what to fill in."""
    ws = wb.create_sheet("Field Guide")
    ws.append(["What to put in every column"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Header colour in each sheet: red = required, amber = conditional,"])
    ws.append(["grey = ignored by CSAP, navy = optional. Hover a header for the same note."])
    ws.append([])

    ws.append(["Sheet", "Column", "Required?", "What it is", "Example"])
    header_row = ws.max_row
    for col in range(1, 6):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for sheet_name, headers in spec.items():
        sheet_guide = guide.get(sheet_name, {})
        if not sheet_guide:
            continue
        for header in headers:
            requirement, description, example = sheet_guide.get(header, ("optional", "", ""))
            label = "reference only" if sheet_name in reference_sheets else requirement
            ws.append([sheet_name, header, label, description, example])
            if label in REQUIREMENT_FILLS:
                ws.cell(row=ws.max_row, column=3).font = Font(bold=True)

    for column, width in zip("ABCDE", (18, 24, 16, 78, 24), strict=True):
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=header_row + 1):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _add_dropdowns(ws: Any, headers: list[str]) -> None:
    """Constrain the columns with a fixed vocabulary so users pick instead of typing."""
    last_row = ws.max_row + SPARE_ROWS
    for index, header in enumerate(headers, start=1):
        options = DROPDOWNS.get(header)
        if not options:
            continue
        validation = DataValidation(
            type="list",
            formula1=options,
            allow_blank=True,
            showDropDown=False,  # openpyxl inverts this: False means "show the arrow"
            errorTitle="Not an allowed value",
            error=f"Choose one of: {options.strip('\"')}",
            promptTitle=header,
            prompt=f"Leave blank to ignore this row, or choose: {options.strip('\"')}"
            if header == "action"
            else f"Choose one of: {options.strip('\"')}",
        )
        ws.add_data_validation(validation)
        column = get_column_letter(index)
        validation.add(f"{column}2:{column}{last_row}")


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


FINDING_HEADERS = ["Severity", "Sheet", "Row", "Column", "What is wrong", "What you should do"]
SEVERITY_FILL = {
    "error": PatternFill("solid", fgColor="FDECEC"),
    "warning": PatternFill("solid", fgColor="FFF5E0"),
    "info": PatternFill("solid", fgColor="E8F1FD"),
}
SEVERITY_FONT = {
    "error": Font(color="A01B1B", bold=True),
    "warning": Font(color="8A5B00", bold=True),
    "info": Font(color="144D9C", bold=True),
}
SEVERITY_ORDER = {"error": 0, "warning": 1, "info": 2}


def build_findings_workbook(
    issues: list[dict[str, Any]],
    filename: str,
    system_name: str,
    counts: dict[str, int],
) -> bytes:
    """One row per finding, with the remediation spelled out, ready to hand to whoever fixes it."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    ws.append([f"Validation findings for {filename}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"System: {system_name}"])
    ws.append(
        [
            f"Errors: {counts.get('errors', 0)}    "
            f"Warnings: {counts.get('warnings', 0)}    "
            f"Planned changes: {counts.get('changes', 0)}"
        ]
    )
    ws.append(["Errors block deployment. Warnings do not, but read them - a warning usually"])
    ws.append(["means a row you edited will be ignored."])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(FINDING_HEADERS)
    for col in range(1, len(FINDING_HEADERS) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    ordered = sorted(
        issues,
        key=lambda i: (
            SEVERITY_ORDER.get(str(i.get("severity")), 9),
            str(i.get("sheet") or ""),
            i.get("row") or 0,
        ),
    )
    for issue in ordered:
        severity = str(issue.get("severity") or "info")
        ws.append(
            [
                severity,
                issue.get("sheet") or "-",
                issue.get("row") or "",
                issue.get("field") or "",
                issue.get("message") or "",
                issue.get("remediation") or "",
            ]
        )
        cell = ws.cell(row=ws.max_row, column=1)
        cell.fill = SEVERITY_FILL.get(severity, SEVERITY_FILL["info"])
        cell.font = SEVERITY_FONT.get(severity, SEVERITY_FONT["info"])

    if not ordered:
        ws.append(["info", "-", "", "", "No issues found.", "Nothing to do - this workbook is ready."])

    for column, width in zip("ABCDEF", (12, 18, 8, 14, 62, 72), strict=True):
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row[4:6]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:F{ws.max_row}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_readme(
    wb: Workbook,
    spec: dict[str, list[str]],
    existing: dict[str, list[dict[str, Any]]],
    product: str,
    product_version: str | None,
    reference_sheets: set[str],
    snapshot_id: str | None = None,
    snapshot_label: str | None = None,
) -> None:
    ws = wb.create_sheet("README")
    ws.append(["Cisco Security Automation Platform - change workbook"])
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([])
    ws.append(["Product", product])
    ws.append(["Detected version", product_version or "unknown"])
    ws.append([PROVENANCE_LABELS["exported"], datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws.append([PROVENANCE_LABELS["snapshot"], snapshot_id or ""])
    ws.append(["Taken from", snapshot_label or ""])
    ws.append([])
    ws.append(["Do not edit the two rows above. CSAP reads them to confirm you are"])
    ws.append(["editing the current configuration and not a stale export."])
    ws.append([])

    _section(ws, "HOW TO USE THIS WORKBOOK")
    for line in (
        "1. Every object already on the device is listed, one row per object.",
        "2. Find the row you want to change, or add a new row at the bottom of the sheet.",
        "3. Set the 'action' cell on that row. Click the cell and pick from the dropdown.",
        "4. Fill in every column whose header is RED - those are mandatory.",
        "5. Save the file and upload it in CSAP under Changes.",
        "6. Fix anything validation reports, then run a dry run before applying.",
        "",
        "A row is ONLY acted on if its 'action' cell is set. Blank means ignore.",
        "That is why you can leave hundreds of rows untouched and change just one.",
        "",
        "HEADER COLOURS",
        "    RED    mandatory - the row is rejected without it",
        "    AMBER  conditional - required only in certain combinations",
        "    NAVY   optional",
        "    GREY   shown for reference, ignored by CSAP",
        "",
        "Hover over any column header for what it means and an example.",
        "The 'Field Guide' tab lists every column in one table.",
    ):
        ws.append([line])

    ws.append([])
    _section(ws, "WHAT EACH ACTION DOES")
    ws.append(["action", "meaning"])
    _header_row(ws, 2)
    for action, meaning in (
        ("create", "Add a new object. The name must not already exist."),
        ("update", "Replace the object with this name. List the full desired value."),
        ("delete", "Remove the object with this name."),
        ("(blank)", "Ignore this row entirely. This is the default."),
    ):
        ws.append([action, meaning])

    ws.append([])
    _section(ws, "EXAMPLES")
    ws.append(["Sheet", "action", "What you type", "Result"])
    _header_row(ws, 4)
    for example in (
        ["Hosts", "create", "name=APP01, value=10.1.1.30", "New host object for one IP address"],
        ["Hosts", "update", "name=WEB01, value=10.1.1.99", "WEB01 now points at 10.1.1.99"],
        ["Hosts", "delete", "name=OLD-SERVER", "OLD-SERVER is removed from the FMC"],
        ["Networks", "create", "name=DMZ-NET, value=10.2.0.0/24", "A subnet. The /24 is required"],
        ["Ranges", "create", "name=POOL, value=10.1.1.10-10.1.1.50", "Contiguous address range"],
        ["Ports", "create", "name=HTTP-ALT, protocol=TCP, port=8080", "TCP service on port 8080"],
        ["Ports", "create", "name=HI-PORTS, protocol=UDP, port=9000-9100", "A port range"],
        ["NetworkGroups", "create", "name=APP-TIER, members=APP01, APP02", "Group of two hosts"],
        ["NetworkGroups", "update", "name=WEB-TIER, members=WEB01, WEB02, WEB03",
         "Replaces the members. List them ALL, not just the new one"],
    ):
        ws.append(example)

    ws.append([])
    _section(ws, "COMMON MISTAKES")
    for line in (
        "Editing a row but leaving 'action' blank        -> nothing happens. Set action=update.",
        "Typing a subnet in the Hosts sheet             -> use Networks. Hosts is one IP only.",
        "Networks value without a prefix (10.2.0.0)     -> add /24 or whatever is correct.",
        "update on a group listing only the new member  -> the others are removed. List all.",
        "Two rows with the same name in one sheet       -> rejected. Each object appears once.",
        "Deleting an object still used by a group       -> remove it from the group first.",
    ):
        ws.append([line])

    ws.append([])
    _section(ws, "SHEETS IN THIS WORKBOOK")
    ws.append(["Sheet", "Rows", "Deployable"])
    _header_row(ws, 3)

    for sheet_name in spec:
        ws.append(
            [
                sheet_name,
                len(existing.get(sheet_name) or []),
                "reference only" if sheet_name in reference_sheets else "yes",
            ]
        )
    ws.append([])
    ws.append(["'reference only' sheets are shown so you can see the configuration,"])
    ws.append(["but CSAP cannot deploy them yet. Make those changes in FMC directly."])

    truncated = [name for name, rows in existing.items() if len(rows) > MAX_ROWS_PER_SHEET]
    if truncated:
        ws.append([])
        ws.append([f"Truncated to {MAX_ROWS_PER_SHEET} rows: {', '.join(truncated)}"])

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 52


def _section(ws: Any, title: str) -> None:
    ws.append([title])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = Font(bold=True, size=12, color="1F3B57")


def _header_row(ws: Any, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
