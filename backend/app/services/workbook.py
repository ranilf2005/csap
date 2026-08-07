"""Reads an uploaded change workbook into plain rows the plugins can validate."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.templates import PROVENANCE_LABELS

logger = logging.getLogger(__name__)

MAX_ROWS_PER_SHEET = 20_000
SKIP_SHEETS = {"README", "INSTRUCTIONS", "NOTES"}


class WorkbookError(ValueError):
    pass


def parse_workbook(path: Path, expected_sheets: dict[str, list[str]]) -> dict[str, list[dict[str, Any]]]:
    """Return {sheet_name: [{column: value}, ...]} for every sheet the plugin knows about."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise WorkbookError(f"could not read the workbook: {exc}") from exc

    try:
        rows: dict[str, list[dict[str, Any]]] = {}
        matched = False

        for sheet_name in wb.sheetnames:
            if sheet_name.upper() in SKIP_SHEETS:
                continue
            canonical = _match_sheet(sheet_name, expected_sheets)
            if canonical is None:
                logger.info("ignoring unknown sheet '%s'", sheet_name)
                continue
            matched = True
            rows[canonical] = _read_sheet(wb[sheet_name], expected_sheets[canonical])

        if not matched:
            raise WorkbookError(
                "no recognised sheets found; expected one of: " + ", ".join(expected_sheets)
            )
        return rows
    finally:
        wb.close()


def read_provenance(path: Path) -> dict[str, str]:
    """Read the export markers CSAP wrote into the README sheet.

    Returns {} for a workbook that did not come from CSAP.
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}

    try:
        if "README" not in wb.sheetnames:
            return {}
        found: dict[str, str] = {}
        for row in wb["README"].iter_rows(values_only=True):
            if not row or row[0] is None or len(row) < 2 or row[1] is None:
                continue
            label = str(row[0]).strip()
            for key, expected in PROVENANCE_LABELS.items():
                if label == expected:
                    found[key] = str(row[1]).strip()
        return found
    except Exception:
        return {}
    finally:
        wb.close()


def _match_sheet(sheet_name: str, expected: dict[str, list[str]]) -> str | None:
    for canonical in expected:
        if canonical.lower() == sheet_name.strip().lower():
            return canonical
    return None


def _read_sheet(ws: Any, headers: list[str]) -> list[dict[str, Any]]:
    iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return []

    # Map the sheet's actual column order onto the plugin's expected headers.
    positions: dict[str, int] = {}
    for position, cell in enumerate(header_row):
        label = str(cell).strip().lower() if cell is not None else ""
        for header in headers:
            if header.lower() == label:
                positions[header] = position

    records: list[dict[str, Any]] = []
    for count, raw in enumerate(iterator):
        if count >= MAX_ROWS_PER_SHEET:
            raise WorkbookError(f"sheet exceeds {MAX_ROWS_PER_SHEET} rows")
        if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        record = {
            header: _clean(raw[pos]) if pos < len(raw) else ""
            for header, pos in positions.items()
        }
        records.append(record)
    return records


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
