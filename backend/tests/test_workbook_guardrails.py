"""Guardrails that stop a second administrator overwriting changes they never saw."""

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.plugins import registry
from app.plugins.base import DiscoveryResult
from app.services.templates import build_workbook
from app.services.workbook import read_provenance

PLUGIN = registry.get("secure_firewall")


def _discovery() -> DiscoveryResult:
    return DiscoveryResult(
        product_version="7.4.1",
        items=[
            {"item_type": "host", "external_id": "h1", "name": "WEB01",
             "payload": {"id": "h1", "name": "WEB01", "type": "Host", "value": "10.1.1.1"}},
        ],
        summary={"host": 1},
    )


def _export(tmp_path: Path, snapshot_id: str) -> Path:
    discovery = _discovery()
    content = build_workbook(
        PLUGIN.template_spec(discovery),
        "secure_firewall",
        "7.4.1",
        existing=PLUGIN.existing_rows(discovery),
        reference_sheets=set(PLUGIN.reference_sheets),
        snapshot_id=snapshot_id,
        snapshot_label="LAB-FMC 2026-08-07",
    )
    path = tmp_path / "export.xlsx"
    path.write_bytes(content)
    return path


def test_export_records_which_snapshot_it_came_from(tmp_path):
    marks = read_provenance(_export(tmp_path, "snap-123"))
    assert marks["snapshot"] == "snap-123"
    assert marks["exported"]


def test_provenance_of_a_foreign_workbook_is_empty(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Hosts"
    wb.active.append(["action", "name", "value"])
    path = tmp_path / "foreign.xlsx"
    wb.save(path)

    assert read_provenance(path) == {}


def test_unreadable_file_does_not_raise(tmp_path):
    path = tmp_path / "junk.xlsx"
    path.write_bytes(b"not a workbook")
    assert read_provenance(path) == {}


def _sheet(path: Path, name: str):
    wb = load_workbook(path, read_only=False, data_only=True)
    return wb[name]


def test_action_column_offers_a_dropdown(tmp_path):
    ws = _sheet(_export(tmp_path, "snap-1"), "Hosts")
    ranges = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]

    assert any("create,update,delete" in f for f in formulas)
    assert any(r.startswith("A2") for r in ranges), "dropdown must start at the first data row"


def test_dropdown_extends_past_the_exported_rows(tmp_path):
    """Users add new objects at the bottom; those cells need the dropdown too."""
    ws = _sheet(_export(tmp_path, "snap-1"), "Hosts")
    action_dv = next(
        dv for dv in ws.data_validations.dataValidation if "create,update,delete" in dv.formula1
    )
    last_row = int(str(action_dv.sqref).split(":")[1][1:])
    assert last_row > ws.max_row


def test_protocol_column_offers_tcp_and_udp(tmp_path):
    ws = _sheet(_export(tmp_path, "snap-1"), "Ports")
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert any("TCP,UDP" in f for f in formulas)


def test_readme_explains_actions_with_examples(tmp_path):
    wb = load_workbook(_export(tmp_path, "snap-1"), read_only=True, data_only=True)
    text = "\n".join(
        str(cell) for row in wb["README"].iter_rows(values_only=True) for cell in row if cell
    )

    assert "HOW TO USE THIS WORKBOOK" in text
    assert "EXAMPLES" in text
    assert "COMMON MISTAKES" in text
    for action in ("create", "update", "delete"):
        assert action in text
    assert "10.2.0.0/24" in text          # a concrete example, not just prose
    assert "List them ALL" in text        # the group-update trap


def test_exported_workbook_still_round_trips(tmp_path):
    """The extra README content must not break parsing."""
    from app.services.workbook import parse_workbook

    discovery = _discovery()
    rows = parse_workbook(_export(tmp_path, "snap-1"), PLUGIN.template_spec(discovery))
    assert PLUGIN.plan(rows, discovery).total == 0


def test_preview_lists_the_rest_calls():
    discovery = _discovery()
    rows = {
        "Hosts": [
            {"action": "create", "name": "APP01", "value": "10.1.1.30"},
            {"action": "update", "name": "WEB01", "value": "10.1.1.99"},
        ]
    }
    calls = PLUGIN.preview(PLUGIN.plan(rows, discovery))

    assert [c["method"] for c in calls] == ["POST", "PUT"]
    assert calls[0]["path"].endswith("/object/hosts")
    assert calls[0]["body"]["value"] == "10.1.1.30"
    assert calls[1]["path"].endswith("/object/hosts/h1")
    assert calls[1]["body"]["id"] == "h1"


def test_preview_of_a_delete_sends_no_body():
    discovery = _discovery()
    calls = PLUGIN.preview(PLUGIN.plan({"Hosts": [{"action": "delete", "name": "WEB01"}]}, discovery))
    assert calls[0]["method"] == "DELETE"
    assert calls[0]["body"] is None


def test_blank_template_also_carries_provenance_and_dropdowns(tmp_path):
    discovery = _discovery()
    content = build_workbook(
        PLUGIN.template_spec(discovery), "secure_firewall", "7.4.1",
        snapshot_id="snap-9", reference_sheets=set(PLUGIN.reference_sheets),
    )
    path = tmp_path / "blank.xlsx"
    path.write_bytes(content)

    assert read_provenance(path)["snapshot"] == "snap-9"
    ws = load_workbook(BytesIO(content))["Hosts"]
    assert any("create,update,delete" in dv.formula1 for dv in ws.data_validations.dataValidation)
