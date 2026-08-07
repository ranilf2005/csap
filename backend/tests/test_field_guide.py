"""The workbook must tell users which columns are mandatory and what to put in them."""

from io import BytesIO

from openpyxl import load_workbook

from app.plugins import registry
from app.plugins.base import DiscoveryResult
from app.services.templates import build_workbook

PLUGIN = registry.get("secure_firewall")


def _discovery() -> DiscoveryResult:
    return DiscoveryResult(
        product_version="7.4.1",
        items=[
            {"item_type": "access_policy", "external_id": "acp1", "name": "Corp-ACP",
             "payload": {"id": "acp1", "name": "Corp-ACP"}},
        ],
        summary={"access_policy": 1},
    )


def _workbook():
    return load_workbook(
        BytesIO(
            build_workbook(
                PLUGIN.template_spec(_discovery()),
                "secure_firewall",
                "7.4.1",
                reference_sheets=set(PLUGIN.reference_sheets),
                guide=PLUGIN.field_guide(),
                snapshot_id="snap-1",
            )
        )
    )


def test_every_deployable_column_is_documented():
    guide = PLUGIN.field_guide()
    for sheet, headers in PLUGIN.template_spec().items():
        if sheet in PLUGIN.reference_sheets:
            continue
        for header in headers:
            assert header in guide.get(sheet, {}), f"{sheet}.{header} has no guidance"


def test_mandatory_headers_are_marked_red():
    ws = _workbook()["Hosts"]
    required = ws.cell(row=1, column=3)   # value
    optional = ws.cell(row=1, column=4)   # description
    assert required.fill.fgColor.rgb.endswith("9C2B2B")
    assert optional.fill.fgColor.rgb.endswith("1F3B57")


def test_headers_carry_a_note_with_requirement_and_example():
    note = _workbook()["Networks"].cell(row=1, column=3).comment.text
    assert "REQUIRED" in note
    assert "prefix length is required" in note
    assert "10.2.0.0/24" in note


def test_field_guide_sheet_lists_every_column():
    ws = _workbook()["Field Guide"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    pairs = {(r[0], r[1]) for r in rows if r[0] and r[1]}

    assert ("Hosts", "value") in pairs
    assert ("AccessRules", "rule_action") in pairs
    assert ("AccessRules", "send_events_to_fmc") in pairs

    lookup = {(r[0], r[1]): r[2] for r in rows if r[0] and r[1]}
    assert lookup[("Hosts", "value")] == "required"
    assert lookup[("Hosts", "description")] == "optional"
    assert lookup[("AccessRules", "send_events_to_fmc")] == "conditional"
    assert lookup[("AccessRules", "applications")] == "not supported"


def test_rule_action_has_a_dropdown_of_real_fmc_actions():
    ws = _workbook()["AccessRules"]
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    action_list = next(f for f in formulas if "ALLOW" in f)
    for action in ("ALLOW", "TRUST", "BLOCK", "MONITOR", "BLOCK_RESET"):
        assert action in action_list


def test_readme_explains_the_header_colours():
    ws = _workbook()["README"]
    text = "\n".join(str(c) for r in ws.iter_rows(values_only=True) for c in r if c)
    assert "RED    mandatory" in text
    assert "Field Guide" in text
