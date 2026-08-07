"""The workbook must show the current configuration, not just empty headers."""

from io import BytesIO

from openpyxl import load_workbook

from app.plugins import registry
from app.plugins.base import DiscoveryResult
from app.services.templates import build_workbook
from app.services.workbook import parse_workbook

PLUGIN = registry.get("secure_firewall")


def _discovery() -> DiscoveryResult:
    return DiscoveryResult(
        product_version="7.4.1",
        items=[
            {"item_type": "host", "external_id": "h1", "name": "WEB01",
             "payload": {"name": "WEB01", "type": "Host", "value": "10.1.1.1", "description": "web"}},
            {"item_type": "network", "external_id": "n1", "name": "DMZ",
             "payload": {"name": "DMZ", "type": "Network", "value": "10.2.0.0/24"}},
            {"item_type": "port", "external_id": "p1", "name": "HTTPS",
             "payload": {"name": "HTTPS", "protocol": "TCP", "port": "443"}},
            {"item_type": "network_group", "external_id": "g1", "name": "WEB-TIER",
             "payload": {"name": "WEB-TIER", "type": "NetworkGroup",
                         "objects": [{"id": "h1", "name": "WEB01", "type": "Host"}],
                         "literals": [{"type": "Network", "value": "10.7.7.0/24"}]}},
            {"item_type": "access_rule", "external_id": "r1", "name": "allow-web",
             "payload": {"name": "allow-web", "action": "ALLOW", "enabled": True,
                         "_policyName": "Corp-ACP",
                         "sourceNetworks": {"objects": [{"name": "DMZ"}]},
                         "destinationNetworks": {"literals": [{"value": "8.8.8.8"}]},
                         "destinationPorts": {"objects": [{"name": "HTTPS"}]},
                         "applications": {"applications": [{"name": "HTTP"}]},
                         "logBegin": True, "logEnd": False}},
            {"item_type": "nat_rule", "external_id": "nr1", "name": "nat-outbound",
             "payload": {"name": "nat-outbound", "natType": "DYNAMIC", "enabled": True,
                         "_policyName": "Corp-NAT",
                         "sourceInterface": {"name": "inside"},
                         "destinationInterface": {"name": "outside"},
                         "originalSource": {"name": "DMZ"},
                         "translatedSource": {"name": "OUTSIDE-IP"}}},
        ],
        summary={"host": 1, "network": 1, "port": 1, "network_group": 1, "access_rule": 1, "nat_rule": 1},
    )


def _sheets(content: bytes) -> dict[str, list[list]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    return {name: [list(r) for r in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}


def _built() -> bytes:
    discovery = _discovery()
    return build_workbook(
        PLUGIN.template_spec(discovery),
        "secure_firewall",
        "7.4.1",
        existing=PLUGIN.existing_rows(discovery),
        reference_sheets=set(PLUGIN.reference_sheets),
    )


def test_objects_are_written_with_a_blank_action():
    rows = _sheets(_built())["Hosts"]
    assert rows[0] == ["action", "name", "value", "description"]
    # openpyxl reads an empty cell back as None; either way the parser treats it as "no action".
    assert not rows[1][0]
    assert rows[1][1:] == ["WEB01", "10.1.1.1", "web"]


def test_blank_template_has_headers_but_no_rows():
    discovery = _discovery()
    blank = build_workbook(PLUGIN.template_spec(discovery), "secure_firewall", "7.4.1")
    sheets = _sheets(blank)

    assert sheets["Hosts"][0] == ["action", "name", "value", "description"]
    for name, rows in sheets.items():
        if name != "README":
            assert len(rows) == 1, f"{name} should contain only a header row"


def test_group_members_are_readable_names_not_ids():
    members = _sheets(_built())["NetworkGroups"][1][2]
    assert "WEB01" in members
    assert "10.7.7.0/24" in members  # literals appear too


def test_access_rules_are_exported_with_flattened_columns():
    header, rule = _sheets(_built())["AccessRules"][:2]
    row = dict(zip(header, rule, strict=True))
    assert row["policy"] == "Corp-ACP"
    assert row["rule_name"] == "allow-web"
    assert row["rule_action"] == "ALLOW"
    assert row["source_networks"] == "DMZ"
    assert row["destination_networks"] == "8.8.8.8"
    assert row["destination_ports"] == "HTTPS"
    assert row["applications"] == "HTTP"
    assert row["enabled"] == "true"


def test_nat_rules_are_exported():
    header, rule = _sheets(_built())["NatRules"][:2]
    row = dict(zip(header, rule, strict=True))
    assert row["policy"] == "Corp-NAT"
    assert row["nat_type"] == "DYNAMIC"
    assert row["source_interface"] == "inside"
    assert row["original_source"] == "DMZ"


def test_readme_reports_row_counts_and_which_sheets_deploy():
    readme = [r for r in _sheets(_built())["README"] if r and r[0]]
    flat = {str(r[0]): [str(c) for c in r[1:] if c is not None] for r in readme}
    assert flat["Hosts"] == ["1", "yes"]
    assert flat["AccessRules"] == ["1", "yes"]
    assert flat["NatRules"] == ["1", "reference only"]


def test_a_downloaded_workbook_round_trips_through_the_parser(tmp_path):
    """Downloading and re-uploading unchanged must produce zero planned changes."""
    path = tmp_path / "round-trip.xlsx"
    path.write_bytes(_built())

    discovery = _discovery()
    rows = parse_workbook(path, PLUGIN.template_spec(discovery))

    assert PLUGIN.validate(rows, discovery).is_valid
    assert PLUGIN.plan(rows, discovery).total == 0


def test_setting_an_action_on_an_exported_row_produces_a_change(tmp_path):
    path = tmp_path / "edited.xlsx"
    path.write_bytes(_built())

    discovery = _discovery()
    rows = parse_workbook(path, PLUGIN.template_spec(discovery))
    rows["Hosts"][0]["action"] = "update"
    rows["Hosts"][0]["value"] = "10.1.1.99"

    plan = PLUGIN.plan(rows, discovery)
    assert plan.total == 1
    assert plan.updates[0]["name"] == "WEB01"
    assert plan.updates[0]["payload"]["value"] == "10.1.1.99"
