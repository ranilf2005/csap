"""Regression tests for findings from the security review."""

import time
from datetime import UTC, datetime, timedelta
from io import BytesIO

import pytest
from fastapi import HTTPException, Request
from fastapi.security import HTTPAuthorizationCredentials
from openpyxl import load_workbook

from app.api.deps import get_current_user
from app.core.security import create_access_token, decode_access_token, waste_a_hash
from app.services.connections import UnsafeTargetError, assert_target_allowed
from app.services.templates import build_findings_workbook, build_workbook


# -- SSRF ------------------------------------------------------------------
@pytest.mark.parametrize(
    "host",
    [
        "127.0.0.1",
        "localhost",
        "169.254.169.254",   # cloud instance metadata
        "0.0.0.0",
        "[::1]",
    ],
)
def test_device_address_cannot_point_back_at_the_platform(host):
    with pytest.raises(UnsafeTargetError):
        assert_target_allowed(host)


@pytest.mark.parametrize("host", ["198.18.1.10", "10.1.1.1", "192.168.1.1", "172.16.5.5"])
def test_real_device_addresses_are_allowed(host):
    assert_target_allowed(host)


def test_a_name_that_does_not_resolve_is_left_to_the_connection_test():
    assert_target_allowed("fmc.invalid.example")


def test_an_empty_address_is_rejected():
    with pytest.raises(UnsafeTargetError):
        assert_target_allowed("   ")


# -- Spreadsheet formula injection -----------------------------------------
DANGEROUS = [
    '=cmd|\'/c calc\'!A0',
    "+1+1",
    "-1+1",
    "@SUM(1)",
]


@pytest.mark.parametrize("payload", DANGEROUS)
def test_exported_device_data_is_never_a_live_formula(payload):
    """An FMC object named '=cmd|...' must not execute when the export is opened."""
    content = build_workbook(
        {"Hosts": ["action", "name", "value", "description"]},
        "secure_firewall",
        "7.4.1",
        existing={"Hosts": [{"action": "", "name": payload, "value": "10.1.1.1"}]},
    )
    ws = load_workbook(BytesIO(content))["Hosts"]
    cell = ws.cell(row=2, column=2)

    assert cell.data_type == "s", "cell was stored as a formula"
    assert cell.value == payload, "the value must survive intact"


@pytest.mark.parametrize("payload", DANGEROUS)
def test_findings_export_is_also_safe(payload):
    issues = [{"severity": "error", "sheet": "Hosts", "row": 2,
               "field": "name", "message": payload, "remediation": payload}]
    content = build_findings_workbook(issues, "c.xlsx", "LAB", {"errors": 1})
    ws = load_workbook(BytesIO(content))["Findings"]

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith(("=", "+", "-", "@")):
                assert cell.data_type == "s"


def test_ordinary_values_are_untouched():
    content = build_workbook(
        {"Hosts": ["action", "name", "value", "description"]},
        "secure_firewall", "7.4.1",
        existing={"Hosts": [{"action": "", "name": "WEB01", "value": "10.1.1.1"}]},
    )
    ws = load_workbook(BytesIO(content))["Hosts"]
    assert ws.cell(row=2, column=2).value == "WEB01"
    assert ws.cell(row=2, column=3).value == "10.1.1.1"


# -- Login timing ----------------------------------------------------------
def test_the_unknown_account_path_still_costs_a_hash():
    """Without this the response time reveals which addresses exist."""
    start = time.perf_counter()
    waste_a_hash()
    elapsed = time.perf_counter() - start
    assert elapsed > 0.005, "the equaliser must be as costly as a real bcrypt check"


# -- Token lifetime --------------------------------------------------------
def test_tokens_carry_an_issue_time_so_they_can_be_invalidated():
    claims = decode_access_token(create_access_token("user-1"))
    assert "iat" in claims
    assert "exp" in claims


# -- Session enforcement ---------------------------------------------------
class _FakeUser:
    def __init__(self, *, must_change_password=False, password_changed_at=None):
        self.id = "user-1"
        self.email = "admin@example.com"
        self.is_active = True
        self.must_change_password = must_change_password
        self.password_changed_at = password_changed_at or datetime(2000, 1, 1, tzinfo=UTC)


class _FakeDb:
    def __init__(self, user):
        self._user = user

    def get(self, _model, _pk):
        return self._user


def _call_deps(user, path="/api/v1/connections"):
    request = Request({"type": "http", "path": path, "headers": [], "query_string": b"",
                       "method": "GET", "scheme": "https", "server": ("csap", 443)})
    creds = HTTPAuthorizationCredentials(
        scheme="Bearer", credentials=create_access_token(user.id)
    )
    return get_current_user(request, creds, _FakeDb(user))


def test_a_valid_session_is_accepted():
    assert _call_deps(_FakeUser()).id == "user-1"


def test_a_token_issued_before_the_password_changed_is_refused():
    """Otherwise a stolen token outlives the password reset meant to revoke it."""
    user = _FakeUser(password_changed_at=datetime.now(UTC) + timedelta(minutes=5))
    with pytest.raises(HTTPException) as exc:
        _call_deps(user)
    assert exc.value.status_code == 401


def test_the_installation_password_cannot_be_used_to_drive_firewalls():
    user = _FakeUser(must_change_password=True)
    with pytest.raises(HTTPException) as exc:
        _call_deps(user)
    assert exc.value.status_code == 403


def test_but_it_can_still_reach_the_change_password_endpoint():
    user = _FakeUser(must_change_password=True)
    assert _call_deps(user, path="/api/v1/auth/change-password").id == "user-1"
