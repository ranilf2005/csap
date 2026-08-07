"""Findings must tell the user what to do, and be downloadable as a spreadsheet."""

from io import BytesIO

from openpyxl import load_workbook

from app.plugins import registry
from app.plugins.base import DiscoveryResult
from app.services.templates import build_findings_workbook

PLUGIN = registry.get("secure_firewall")


def _discovery() -> DiscoveryResult:
    return DiscoveryResult(
        product_version="7.4.1",
        items=[
            {"item_type": "host", "external_id": "h1", "name": "WEB01",
             "payload": {"id": "h1", "name": "WEB01", "type": "Host", "value": "10.1.1.1"}},
        ],
        summary={"host": 1},
    )


def _issues(rows):
    return PLUGIN.validate(rows, _discovery()).issues


def test_every_finding_carries_remediation():
    rows = {
        "Hosts": [
            {"action": "create", "name": "WEB01", "value": "10.1.1.1"},      # already exists
            {"action": "create", "name": "BAD", "value": "not-an-ip"},       # bad address
            {"action": "update", "name": "GHOST", "value": "10.1.1.5"},      # missing
            {"action": "sideways", "name": "X", "value": "10.1.1.6"},        # bad action
        ],
        "Networks": [{"action": "create", "name": "N1", "value": "10.2.0.0"}],   # no prefix
        "Ports": [{"action": "create", "name": "P1", "protocol": "SCTP", "port": "99999"}],
        "NetworkGroups": [{"action": "create", "name": "G1", "members": "NOPE"}],
    }
    issues = _issues(rows)
    assert issues
    for issue in issues:
        assert issue.remediation.strip(), f"no remediation for: {issue.message}"


def test_remediation_is_specific_not_generic():
    rows = {"Networks": [{"action": "create", "name": "N1", "value": "10.2.0.0"}]}
    fix = next(i.remediation for i in _issues(rows) if "prefix" in i.message)
    assert "10.2.0.0/24" in fix


def test_host_bits_warning_names_the_correct_network():
    rows = {"Networks": [{"action": "create", "name": "N1", "value": "10.2.0.5/24"}]}
    issue = next(i for i in _issues(rows) if i.severity == "warning")
    assert "10.2.0.0/24" in issue.remediation


def test_reference_sheet_warning_does_not_sound_like_a_failure():
    rows = {"AccessRules": [{"action": "create", "policy": "ACP", "rule_name": "r1"}]}
    issue = next(i for i in _issues(rows) if i.sheet == "AccessRules")
    assert issue.severity == "warning"
    assert "reference only" in issue.message
    assert "No action needed" in issue.remediation


def _sheet(content: bytes):
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    return [list(r) for r in wb["Findings"].iter_rows(values_only=True)]


def test_findings_workbook_lists_severity_message_and_fix():
    rows = {"Hosts": [{"action": "create", "name": "BAD", "value": "not-an-ip"}]}
    issues = [
        {"severity": i.severity, "sheet": i.sheet, "row": i.row,
         "field": i.field, "message": i.message, "remediation": i.remediation}
        for i in _issues(rows)
    ]

    content = build_findings_workbook(issues, "changes.xlsx", "LAB-FMC",
                                      {"errors": 1, "warnings": 0, "changes": 0})
    grid = _sheet(content)

    header = next(r for r in grid if r and r[0] == "Severity")
    assert header[:6] == ["Severity", "Sheet", "Row", "Column", "What is wrong", "What you should do"]

    finding = grid[grid.index(header) + 1]
    assert finding[0] == "error"
    assert finding[1] == "Hosts"
    assert "not a valid IP address" in finding[4]
    assert finding[5]


def test_errors_are_listed_before_warnings():
    issues = [
        {"severity": "warning", "sheet": "Hosts", "row": 5, "field": None,
         "message": "w", "remediation": "x"},
        {"severity": "error", "sheet": "Hosts", "row": 9, "field": None,
         "message": "e", "remediation": "y"},
    ]
    grid = _sheet(build_findings_workbook(issues, "c.xlsx", "LAB", {}))
    header_index = next(i for i, r in enumerate(grid) if r and r[0] == "Severity")
    assert [r[0] for r in grid[header_index + 1:]] == ["error", "warning"]


def test_clean_workbook_still_produces_a_usable_file():
    grid = _sheet(build_findings_workbook([], "c.xlsx", "LAB", {"errors": 0}))
    assert any("No issues found." in str(cell) for row in grid for cell in row if cell)
